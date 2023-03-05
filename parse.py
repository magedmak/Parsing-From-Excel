# File: parse.py
#
# Author:   Maged Magdy Asaad
# Date:     February, 14, 2023
#
# This work was developed under the supervision of Valeo Testing Academy

import openpyxl as xl
import xml.etree.ElementTree as ET
import os

try:
    # reading the Excel file
    wb = xl.load_workbook('Generationfile.xlsx')
    ws = wb['Sheet1']

    try:
        # reading the xml file
        tree = ET.parse('CanComm_RX_[ESP_HL_Radgeschw_02].xml')
        root = tree.getroot()

        # create the result directory if it doesn't exist
        if not os.path.exists("result"):
            os.mkdir("result")

        # replace data
        for i in range(5):  # 5 Modified copies of xml file
            j = 0
            while True:
                if ws.cell(row=2+i, column=3+j).value is None:  # if column is empty, exit loop
                    break

                # search for text in xml file and replace it with Excel data
                for element in root.iter():
                    if element.text == str(ws.cell(row=2+i, column=3+j).value):
                        element.text = str(ws.cell(row=2+i, column=3+j+1).value)
                j += 2

            try:
                # save copy of the xml file with the name in Excel sheet
                tree.write("result/" + ws.cell(row=2+i, column=2).value + ".xml")
            except TypeError:
                print("Something went wrong when saving the XML file")
            finally:
                print(f"copy {i+1} has been saved!")

    except ImportError:
        print("Something went wrong when opening the XML file")
except ImportError:
    print("Something went wrong when opening the Excel file")
